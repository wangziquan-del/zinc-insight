from __future__ import annotations

import argparse
import ast
import hashlib
import json
import math
import os
import re
import shutil
import statistics
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import xlrd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ZINC_DIR = Path(r"D:\拷贝文件\E\永安\锌")
INGOT_BOOK = ZINC_DIR / "【永安期货研究中心】锌数据库-锭篇.xlsx"
DEMAND_BOOK = ZINC_DIR / "【永安期货研究中心】锌数据库 - 需求篇(1).xlsx"
STONEX_BOOK = ZINC_DIR / "zinc-sto-data-tables-may-2026.xls"
COMPANY_BOOK = ZINC_DIR / "全球锌企季度产量梳理.xlsx"
TIN_REFERENCE = Path(r"C:\Users\Yitian Shen\tin-insight-site\index.html")
LOCAL_MCP_SERVER = Path(r"D:\拷贝文件\E\永安\2025 永安\zhiji_mcp_server.py")

SERIES_URL = "https://zhiji-ai.xyz/commodity/api/series"
QUOTE_URL = "https://zhiji-ai.xyz/guan/api/quote"
KLINE_URL = "https://zhiji-ai.xyz/guan/api/kline"

COLORS = ["#f4b942", "#3dd6b6", "#6e9fff", "#ff758f", "#a989ff", "#ff9f43"]
INDICATORS = {
    "lme_price": "FU00016158",
    "shfe_stock": "ID00188293",
    "lme_stock": "FU00016163",
    "social_stock": "ID00188329",
    "concentrate_port_stock": "ID00188330",
    "tc_import": "ID00408213",
    "tc_north": "ID00408211",
    "tc_south": "ID00408212",
    "refined_output": "ID01510883",
    "concentrate_output": "ID01001563",
    "concentrate_import": "a10001843",
    "galvanized_rate": "ID00366835",
    "zinc_oxide_rate": "ID01002075",
    "die_cast_rate": "ID01002076",
    "shanghai_premium": "ID02038785",
    "refined_import": "CM0000138625",
    "refined_export": "CM0000138627",
    "galvanized_inventory": "ID00366838",
    "galvanized_mill_inventory": "ID00366837",
}


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        output = float(value)
    except (TypeError, ValueError):
        return None
    return output if math.isfinite(output) else None


def clean_number(value: Any, digits: int = 6) -> float | None:
    parsed = number(value)
    return round(parsed, digits) if parsed is not None else None


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    match = re.match(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if not match:
        return None
    try:
        return date(*(int(part) for part in match.groups())).isoformat()
    except ValueError:
        return None


def display_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return iso_date(value)
    if isinstance(value, float):
        return round(value, 6)
    if value is None:
        return ""
    return value


def extract_xy(
    worksheet,
    date_col: int,
    value_col: int,
    min_row: int = 1,
    scale: float = 1.0,
    start: str = "2021-01-01",
) -> list[list[Any]]:
    points: dict[str, float] = {}
    max_col = max(date_col, value_col)
    for row in worksheet.iter_rows(min_row=min_row, max_col=max_col, values_only=True):
        day = iso_date(row[date_col - 1])
        value = number(row[value_col - 1])
        if day and day >= start and value is not None:
            points[day] = value * scale
    return [[day, clean_number(value)] for day, value in sorted(points.items())]


def series_latest(series: list[list[Any]]) -> list[Any]:
    return series[-1] if series else [None, None]


def seasonal_chart(series: list[list[Any]], years: int = 5) -> dict[str, Any]:
    if not series:
        return {"labels": [], "datasets": []}
    available = sorted({int(item[0][:4]) for item in series if item[0] >= "2021-01-01"})[-years:]
    yearly_counts = {
        year: sum(1 for item in series if int(item[0][:4]) == year)
        for year in available
    }
    monthly = bool(yearly_counts) and max(yearly_counts.values()) <= 15
    label_for = (lambda day: day[5:7]) if monthly else (lambda day: day[5:])
    labels = sorted({label_for(item[0]) for item in series if int(item[0][:4]) in available})
    datasets = []
    maximum = max(available)
    for index, year in enumerate(available):
        values = {label_for(item[0]): item[1] for item in series if int(item[0][:4]) == year}
        color = COLORS[index % len(COLORS)]
        datasets.append(
            {
                "label": str(year),
                "data": [values.get(label) for label in labels],
                "borderColor": color,
                "backgroundColor": color + "18",
                "borderWidth": 2.8 if year == maximum else 1.35,
            }
        )
    return {"labels": labels, "datasets": datasets}


def read_seasonal_matrix(
    worksheet,
    *,
    header_row: int,
    date_col: int,
    first_year_col: int,
    last_year_col: int,
) -> list[list[Any]]:
    years: dict[int, int] = {}
    for column in range(first_year_col, last_year_col + 1):
        match = re.search(r"(20\d{2})", str(worksheet.cell(header_row, column).value or ""))
        if match:
            years[column] = int(match.group(1))
    points: dict[str, float] = {}
    for row in range(header_row + 1, worksheet.max_row + 1):
        month_day = str(worksheet.cell(row, date_col).value or "").strip()
        if not re.fullmatch(r"\d{2}-\d{2}", month_day):
            continue
        for column, year in years.items():
            value = number(worksheet.cell(row, column).value)
            if value is None:
                continue
            day = iso_date(f"{year}-{month_day}")
            if day:
                points[day] = value
    return [[day, clean_number(value)] for day, value in sorted(points.items())]


def monthly_difference(
    left: list[list[Any]], right: list[list[Any]], scale: float = 1.0
) -> list[list[Any]]:
    left_monthly = month_end(left)
    right_monthly = month_end(right)
    return [
        [f"{month}-01", clean_number((left_monthly[month] - right_monthly[month]) * scale)]
        for month in sorted(set(left_monthly) & set(right_monthly))
    ]


def calendar_month_percentile(series: list[list[Any]]) -> float | None:
    if not series:
        return None
    latest_day, latest_value = series[-1]
    latest_year, latest_month = int(latest_day[:4]), latest_day[5:7]
    peers = [
        value
        for day, value in series
        if int(day[:4]) < latest_year and day[5:7] == latest_month
    ]
    if not peers:
        return None
    return clean_number(sum(value <= latest_value for value in peers) / len(peers) * 100, 1)


def continuous_chart(
    series_map: dict[str, list[list[Any]]], limit: int | None = 360
) -> dict[str, Any]:
    labels = sorted({item[0] for series in series_map.values() for item in series})
    if limit:
        labels = labels[-limit:]
    datasets = []
    for index, (label, series) in enumerate(series_map.items()):
        lookup = {item[0]: item[1] for item in series}
        color = COLORS[index % len(COLORS)]
        datasets.append(
            {
                "label": label,
                "data": [lookup.get(day) for day in labels],
                "borderColor": color,
                "backgroundColor": color + "18",
                "borderWidth": 2.1,
            }
        )
    return {"labels": labels, "datasets": datasets}


def month_end(series: list[list[Any]]) -> dict[str, float]:
    output: dict[str, tuple[str, float]] = {}
    for day, value in series:
        month = day[:7]
        if month not in output or day > output[month][0]:
            output[month] = (day, value)
    return {month: item[1] for month, item in output.items()}


def load_api_key() -> str:
    configured = os.environ.get("ZHIJI_API_KEY", "").strip()
    if configured:
        return configured
    if not LOCAL_MCP_SERVER.exists():
        return ""
    try:
        tree = ast.parse(LOCAL_MCP_SERVER.read_text(encoding="utf-8"))
        for node in tree.body:
            if not isinstance(node, (ast.Assign, ast.AnnAssign)):
                continue
            targets = node.targets if isinstance(node, ast.Assign) else [node.target]
            if not any(isinstance(target, ast.Name) and target.id == "API_KEY" for target in targets):
                continue
            value_node = node.value
            if isinstance(value_node, ast.Constant) and isinstance(value_node.value, str):
                return value_node.value.strip()
    except Exception:
        return ""
    return ""


def request_json(url: str, params: dict[str, Any], timeout: int = 45) -> dict[str, Any]:
    request = Request(
        url + "?" + urlencode(params),
        headers={"User-Agent": "Zinc Insight Builder/1.0"},
    )
    with urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def fetch_series(key: str, indicator_id: str) -> tuple[list[list[Any]], dict[str, Any]]:
    payload = request_json(
        SERIES_URL,
        {
            "id": indicator_id,
            "start": "2021-01-01",
            "end": date.today().isoformat(),
            "key": key,
        },
    )
    if payload.get("error"):
        raise RuntimeError(str(payload["error"]))
    points: dict[str, float] = {}
    for item in payload.get("points") or []:
        day = iso_date(item.get("date"))
        value = number(item.get("value"))
        if day and value is not None:
            points[day] = value
    series = [[day, clean_number(value)] for day, value in sorted(points.items())]
    meta = {
        "id": indicator_id,
        "source": "直集 API",
        "name": payload.get("name"),
        "unit": payload.get("unit"),
        "frequency": payload.get("frequency"),
        "organization": payload.get("src_org"),
        "dataLatest": payload.get("data_latest") or (series[-1][0] if series else None),
    }
    return series, meta


def fetch_api_bundle(key: str) -> tuple[dict[str, list[list[Any]]], dict[str, Any]]:
    if not key:
        return {}, {}
    series: dict[str, list[list[Any]]] = {}
    meta: dict[str, Any] = {}
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {
            pool.submit(fetch_series, key, indicator_id): label
            for label, indicator_id in INDICATORS.items()
        }
        for future in as_completed(futures):
            label = futures[future]
            try:
                series[label], meta[label] = future.result()
            except Exception as exc:
                series[label] = []
                meta[label] = {
                    "id": INDICATORS[label],
                    "source": "直集 API",
                    "error": type(exc).__name__,
                }
    return series, meta


def fetch_quote(key: str) -> dict[str, Any]:
    if not key:
        return {}
    payload = request_json(QUOTE_URL, {"symbols": "ZN", "key": key}, timeout=30)
    rows = payload.get("quotes") or payload.get("data") or []
    item = next((row for row in rows if not row.get("error")), {})
    if not item:
        return {}
    return {
        "symbol": item.get("symbol") or "ZN0",
        "name": item.get("name") or "沪锌",
        "last": clean_number(item.get("last")),
        "changePct": clean_number(item.get("change_pct")),
        "open": clean_number(item.get("open")),
        "high": clean_number(item.get("high")),
        "low": clean_number(item.get("low")),
        "volume": clean_number(item.get("volume")),
        "openInterest": clean_number(item.get("open_interest")),
        "asOf": (str(item.get("date") or "") + " " + str(item.get("time") or "")).strip(),
        "source": "直集 API 实时行情",
    }


def fetch_kline(key: str, limit: int = 360) -> list[dict[str, Any]]:
    if not key:
        return []
    payload = request_json(
        KLINE_URL,
        {"symbol": "ZN", "freq": "D", "cont": 1, "limit": limit, "key": key},
        timeout=45,
    )
    rows = payload.get("bars") or payload.get("data") or []
    output = []
    for item in rows:
        stamp = str(
            item.get("time")
            or item.get("date")
            or item.get("datetime")
            or item.get("trade_date")
            or ""
        )
        opened = number(item.get("open", item.get("o")))
        high = number(item.get("high", item.get("h")))
        low = number(item.get("low", item.get("l")))
        close = number(item.get("close", item.get("c")))
        if not stamp or None in (opened, high, low, close):
            continue
        output.append(
            {
                "time": stamp,
                "o": clean_number(opened),
                "h": clean_number(high),
                "l": clean_number(low),
                "c": clean_number(close),
                "v": clean_number(item.get("volume", item.get("v"))),
            }
        )
    output.sort(key=lambda item: item["time"])
    return output[-limit:]


def moving_average(values: list[float], period: int) -> list[float | None]:
    output: list[float | None] = [None] * len(values)
    if period <= 0:
        return output
    running = 0.0
    for index, value in enumerate(values):
        running += value
        if index >= period:
            running -= values[index - period]
        if index >= period - 1:
            output[index] = round(running / period, 4)
    return output


def calculate_rsi(values: list[float], period: int = 14) -> float | None:
    if len(values) <= period:
        return None
    changes = [values[index] - values[index - 1] for index in range(1, len(values))]
    gains = [max(change, 0) for change in changes[-period:]]
    losses = [max(-change, 0) for change in changes[-period:]]
    avg_gain = statistics.fmean(gains)
    avg_loss = statistics.fmean(losses)
    if avg_loss == 0:
        return 100.0
    return round(100 - 100 / (1 + avg_gain / avg_loss), 2)


def enrich_kline(rows: list[dict[str, Any]]) -> dict[str, Any]:
    closes = [float(row["c"]) for row in rows]
    mas = {
        "MA5": moving_average(closes, 5),
        "MA20": moving_average(closes, 20),
        "MA60": moving_average(closes, 60),
        "MA120": moving_average(closes, 120),
    }
    if not rows:
        return {"candles": [], "mas": mas, "rsi14": None}
    latest = closes[-1]
    ma20 = mas["MA20"][-1]
    ma60 = mas["MA60"][-1]
    recent = closes[-60:]
    rsi = calculate_rsi(closes)
    trend = "多头结构" if ma20 and ma60 and latest > ma20 > ma60 else (
        "空头结构" if ma20 and ma60 and latest < ma20 < ma60 else "震荡结构"
    )
    return {
        "candles": rows,
        "mas": mas,
        "rsi14": rsi,
        "latest": latest,
        "high60": max(recent) if recent else None,
        "low60": min(recent) if recent else None,
        "trend": trend,
    }


def fallback_kline(price_series: list[list[Any]], limit: int = 360) -> list[dict[str, Any]]:
    rows = []
    previous = None
    for day, value in price_series[-limit:]:
        opened = previous if previous is not None else value
        rows.append({"time": day, "o": opened, "h": max(opened, value), "l": min(opened, value), "c": value, "v": None})
        previous = value
    return rows


def read_excel_sources() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    warnings: list[dict[str, Any]] = []
    for path in (INGOT_BOOK, DEMAND_BOOK, STONEX_BOOK, COMPANY_BOOK):
        if not path.exists():
            warnings.append({"source": path.name, "issue": "文件未找到"})
    if warnings:
        missing = ", ".join(item["source"] for item in warnings)
        raise FileNotFoundError("Missing required zinc sources: " + missing)

    ingot_values = load_workbook(INGOT_BOOK, read_only=True, data_only=True)
    demand_values = load_workbook(DEMAND_BOOK, read_only=True, data_only=True)
    company_values = load_workbook(COMPANY_BOOK, read_only=True, data_only=True)

    local: dict[str, Any] = {
        "price": extract_xy(ingot_values["沪锌价格"], 1, 2, min_row=3),
        "month_spread": read_seasonal_matrix(
            ingot_values["月差"], header_row=5, date_col=1, first_year_col=2, last_year_col=8
        ),
        "premium_shanghai": extract_xy(ingot_values["升贴水数据"], 11, 12),
        "premium_guangdong": extract_xy(ingot_values["升贴水数据"], 11, 13),
        "premium_tianjin": extract_xy(ingot_values["升贴水数据"], 11, 14),
        "tc_domestic": extract_xy(ingot_values["TC"], 1, 2),
        "tc_import": extract_xy(ingot_values["TC"], 1, 3),
        "smelting_profit_domestic": extract_xy(ingot_values["利润"], 16, 17),
        "smelting_profit_import": extract_xy(ingot_values["利润"], 30, 31),
        "ore_import": extract_xy(ingot_values["锌矿进口"], 11, 12, scale=0.0001),
        "pmi_china": extract_xy(demand_values["PMI"], 1, 2),
        "pmi_us": extract_xy(demand_values["PMI"], 11, 12),
    }

    local["mine_balance"] = read_mine_balance(ingot_values["锌矿平衡"])
    local["refined_balance"] = read_refined_balance(ingot_values["锌锭平衡"])
    local["projects"] = read_projects(ingot_values["海外新开工"])
    local["companies"] = {
        "mine": read_table(company_values["锌矿企业·季度产量"]),
        "smelter": read_table(company_values["锌锭冶炼企业·季度产量"]),
        "capex": read_table(company_values["锌矿企业·资本开支"]),
        "log": read_table(company_values["更新日志"]),
    }

    ingot_values.close()
    demand_values.close()
    company_values.close()
    local["stonex"] = read_stonex()
    return local, warnings


def read_mine_balance(worksheet) -> list[dict[str, Any]]:
    rows = []
    for row in worksheet.iter_rows(min_row=2, max_col=17, values_only=True):
        day = iso_date(row[0])
        production = number(row[1])
        total_supply = number(row[10])
        consumption = number(row[14])
        balance = number(row[15])
        if not day or production is None or total_supply is None:
            continue
        rows.append(
            {
                "date": day[:7],
                "production": clean_number(production),
                "importPhysical": clean_number(row[4]),
                "importMetal": clean_number(row[7]),
                "totalSupply": clean_number(total_supply),
                "smelterUse": clean_number(consumption),
                "balance": clean_number(balance),
                "cumulative": clean_number(row[16]),
                "basis": "锭篇·锌矿平衡；进口矿按45%折金属量，锌锭产量按表内系数折算矿耗",
            }
        )
    return rows


def read_refined_balance(worksheet) -> list[dict[str, Any]]:
    rows = []
    for row in worksheet.iter_rows(min_row=3, max_col=26, values_only=True):
        day = iso_date(row[1])
        production = number(row[2])
        supply = number(row[14])
        apparent = number(row[21])
        if not day or production is None or supply is None:
            continue
        rows.append(
            {
                "date": day[:7],
                "production": clean_number(production),
                "netImport": clean_number(row[11]),
                "totalSupply": clean_number(supply),
                "socialStock": clean_number(row[17]),
                "stockChange": clean_number(row[18]),
                "bondedStock": clean_number(row[19]),
                "bondedChange": clean_number(row[20]),
                "apparentConsumption": clean_number(apparent),
                "basis": "锭篇·锌锭平衡；表观消费由总供应扣除社会及保税库存变化倒算",
            }
        )
    return rows


def read_projects(worksheet) -> list[dict[str, Any]]:
    rows = []
    for row in worksheet.iter_rows(min_row=3, max_col=7, values_only=True):
        country, project, stage, value_2024, value_2025, value_2026, note = row
        if not project:
            continue
        rows.append(
            {
                "country": display_value(country),
                "project": display_value(project),
                "stage": display_value(stage),
                "y2024": clean_number(value_2024) if number(value_2024) is not None else display_value(value_2024),
                "y2025": clean_number(value_2025) if number(value_2025) is not None else display_value(value_2025),
                "y2026": clean_number(value_2026) if number(value_2026) is not None else display_value(value_2026),
                "note": display_value(note),
            }
        )
    return rows


def read_table(worksheet) -> dict[str, Any]:
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        values = [display_value(value) for value in row]
        while values and values[-1] in ("", None):
            values.pop()
        if values:
            rows.append(values)
    return {"headers": rows[0] if rows else [], "rows": rows[1:] if rows else []}


def xlrd_value(sheet, row: int, column: int) -> Any:
    value = sheet.cell_value(row, column)
    return display_value(value)


def read_stonex() -> dict[str, Any]:
    book = xlrd.open_workbook(str(STONEX_BOOK), on_demand=True)
    output: dict[str, Any] = {}

    key_sheet = book.sheet_by_name("KeyForecasts")
    years = [int(number(key_sheet.cell_value(6, col)) or 0) for col in (2, 4, 6, 8)]
    forecast = []
    for row in range(8, min(key_sheet.nrows, 21)):
        label = str(key_sheet.cell_value(row, 1) or "").strip()
        if not label:
            continue
        values = [clean_number(key_sheet.cell_value(row, col)) for col in (2, 4, 6, 8)]
        forecast.append({"metric": label, "values": dict(zip(map(str, years), values))})
    output["forecast"] = {"years": years, "rows": forecast}

    quarter_sheet = book.sheet_by_name("QuarterlyBalance")
    quarterly = []
    active_year = None
    for row in range(4, quarter_sheet.nrows):
        raw_year = number(quarter_sheet.cell_value(row, 1))
        if raw_year:
            active_year = int(raw_year)
        quarter = str(quarter_sheet.cell_value(row, 2) or "").strip()
        if not active_year or not quarter or quarter == "Total/Avg":
            continue
        refined_supply = number(quarter_sheet.cell_value(row, 3))
        if refined_supply is None:
            continue
        quarterly.append(
            {
                "period": f"{active_year}{quarter}",
                "supply": clean_number(refined_supply),
                "consumption": clean_number(quarter_sheet.cell_value(row, 4)),
                "balance": clean_number(quarter_sheet.cell_value(row, 5)),
                "stockDays": clean_number(quarter_sheet.cell_value(row, 7)),
                "lmeCash": clean_number(quarter_sheet.cell_value(row, 9)),
            }
        )
    output["quarterly"] = quarterly

    output["consumption"] = read_stonex_country_sheet(book, "Consumption", header_row=6, start_row=8)
    output["mineProduction"] = read_stonex_country_sheet(book, "MineProduction", header_row=7, start_row=9)
    output["smelterProduction"] = read_stonex_country_sheet(book, "SmelterProduction", header_row=7, start_row=9)
    output["globalBalance"] = read_stonex_metric_sheet(book, "GlobalBalance", header_row=6, start_row=8)
    output["chinaTrade"] = read_stonex_metric_sheet(book, "ChinaTrade", header_row=7, start_row=8)
    book.release_resources()
    return output


def read_stonex_country_sheet(
    book, sheet_name: str, header_row: int, start_row: int
) -> dict[str, Any]:
    sheet = book.sheet_by_name(sheet_name)
    years = []
    columns = []
    for column in range(2, sheet.ncols):
        parsed = number(sheet.cell_value(header_row, column))
        if parsed and 2000 <= parsed <= 2100:
            years.append(int(parsed))
            columns.append(column)
    rows = []
    for row in range(start_row, sheet.nrows):
        name = str(sheet.cell_value(row, 1) or "").strip()
        if not name or "Growth Rate" in name or "% change" in name:
            continue
        values = [clean_number(sheet.cell_value(row, column)) for column in columns]
        if not any(value is not None for value in values):
            continue
        rows.append({"name": name, "values": dict(zip(map(str, years), values))})
    return {"years": years, "rows": rows}


def read_stonex_metric_sheet(
    book, sheet_name: str, header_row: int, start_row: int
) -> dict[str, Any]:
    sheet = book.sheet_by_name(sheet_name)
    years = []
    columns = []
    for column in range(2, sheet.ncols):
        parsed = number(sheet.cell_value(header_row, column))
        if parsed and 2000 <= parsed <= 2100:
            years.append(int(parsed))
            columns.append(column)
    rows = []
    for row in range(start_row, sheet.nrows):
        name = str(sheet.cell_value(row, 1) or "").strip()
        if not name:
            continue
        values = [clean_number(sheet.cell_value(row, column)) for column in columns]
        if any(value is not None for value in values):
            rows.append({"metric": name, "values": dict(zip(map(str, years), values))})
    return {"years": years, "rows": rows}


def extend_refined_balance(
    historical: list[dict[str, Any]], api_series: dict[str, list[list[Any]]]
) -> list[dict[str, Any]]:
    production = month_end(api_series.get("refined_output", []))
    imports = month_end(api_series.get("refined_import", []))
    exports = month_end(api_series.get("refined_export", []))
    stocks = month_end(api_series.get("social_stock", []))
    previous_stock = historical[-1].get("socialStock") if historical else None
    existing = {row["date"] for row in historical}
    output = list(historical)
    for month in sorted(production):
        if month in existing or month < "2026-01":
            continue
        prod = production[month]
        net_import = (imports.get(month, 0) - exports.get(month, 0)) / 10000
        supply = prod + net_import
        current_stock = stocks.get(month)
        stock_change = (
            current_stock - previous_stock
            if current_stock is not None and previous_stock is not None
            else None
        )
        apparent = supply - stock_change if stock_change is not None else None
        output.append(
            {
                "date": month,
                "production": clean_number(prod),
                "netImport": clean_number(net_import),
                "totalSupply": clean_number(supply),
                "socialStock": clean_number(current_stock),
                "stockChange": clean_number(stock_change),
                "bondedStock": None,
                "bondedChange": None,
                "apparentConsumption": clean_number(apparent),
                "basis": "直集 API 续接；净进口=精炼锌进口-出口，表消未含锌合金净进口与保税库存变化",
            }
        )
        if current_stock is not None:
            previous_stock = current_stock
    return output


def extend_mine_balance(
    historical: list[dict[str, Any]],
    api_series: dict[str, list[list[Any]]],
    ore_import_fallback: list[list[Any]],
) -> list[dict[str, Any]]:
    production = month_end(api_series.get("concentrate_output", []))
    imports = month_end(ore_import_fallback)
    refined = month_end(api_series.get("refined_output", []))
    existing = {row["date"] for row in historical}
    output = list(historical)
    for month in sorted(production):
        if month in existing or month < "2026-01":
            continue
        prod = production[month]
        import_physical = imports.get(month)
        import_metal = import_physical * 0.45 if import_physical is not None else None
        total_supply = prod + import_metal if import_metal is not None else None
        smelter_use = refined.get(month)
        smelter_use = smelter_use * 0.88 if smelter_use is not None else None
        balance = total_supply - smelter_use if total_supply is not None and smelter_use is not None else None
        output.append(
            {
                "date": month,
                "production": clean_number(prod),
                "importPhysical": clean_number(import_physical),
                "importMetal": clean_number(import_metal),
                "totalSupply": clean_number(total_supply),
                "smelterUse": clean_number(smelter_use),
                "balance": clean_number(balance),
                "cumulative": None,
                "basis": "直集/API与锭篇进口续接；进口矿按45%折金属量，矿耗按精炼锌产量×0.88估算",
            }
        )
    return output


def metric_lookup(rows: list[dict[str, Any]], *needles: str) -> dict[str, Any]:
    for row in rows:
        label = str(row.get("metric") or row.get("name") or "").lower()
        if all(needle.lower() in label for needle in needles):
            return row
    return {}


def source_file_meta(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        "size": stat.st_size,
    }


def build_data() -> dict[str, Any]:
    local, warnings = read_excel_sources()
    key = load_api_key()
    api_series, api_meta = fetch_api_bundle(key)

    series = {
        "price": local["price"],
        "lme_price": api_series.get("lme_price", []),
        "shfe_stock": api_series.get("shfe_stock", []),
        "lme_stock": api_series.get("lme_stock", []),
        "social_stock": api_series.get("social_stock", []),
        "concentrate_port_stock": api_series.get("concentrate_port_stock", []),
        "tc_import": api_series.get("tc_import") or local["tc_import"],
        "tc_north": api_series.get("tc_north") or local["tc_domestic"],
        "tc_south": api_series.get("tc_south", []),
        "refined_output": api_series.get("refined_output", []),
        "concentrate_output": api_series.get("concentrate_output", []),
        "concentrate_import": api_series.get("concentrate_import") or local["ore_import"],
        "galvanized_rate": api_series.get("galvanized_rate", []),
        "zinc_oxide_rate": api_series.get("zinc_oxide_rate", []),
        "die_cast_rate": api_series.get("die_cast_rate", []),
        "shanghai_premium": api_series.get("shanghai_premium") or local["premium_shanghai"],
        "refined_import": api_series.get("refined_import", []),
        "refined_export": api_series.get("refined_export", []),
        "galvanized_inventory": api_series.get("galvanized_inventory", []),
        "galvanized_mill_inventory": api_series.get("galvanized_mill_inventory", []),
    }
    refined_net_import = monthly_difference(
        series["refined_import"], series["refined_export"], scale=0.0001
    )

    try:
        quote = fetch_quote(key)
    except Exception as exc:
        warnings.append({"source": "直集实时行情", "issue": type(exc).__name__})
        quote = {}
    try:
        kline_rows = fetch_kline(key)
    except Exception as exc:
        warnings.append({"source": "直集日K", "issue": type(exc).__name__})
        kline_rows = []
    if not kline_rows:
        kline_rows = fallback_kline(series["price"])
    kline = enrich_kline(kline_rows)

    if quote.get("last") is None and kline.get("latest") is not None:
        quote = {
            "symbol": "ZN",
            "name": "沪锌",
            "last": kline["latest"],
            "changePct": None,
            "asOf": kline_rows[-1]["time"],
            "source": "直集日K / Excel 回退",
        }

    refined_balance = extend_refined_balance(local["refined_balance"], api_series)
    mine_balance = extend_mine_balance(local["mine_balance"], api_series, series["concentrate_import"])
    latest = {
        "shfe": [quote.get("asOf"), quote.get("last")],
        "lme": series_latest(series["lme_price"]),
        "shfeStock": series_latest(series["shfe_stock"]),
        "lmeStock": series_latest(series["lme_stock"]),
        "socialStock": series_latest(series["social_stock"]),
        "concentratePortStock": series_latest(series["concentrate_port_stock"]),
        "tcImport": series_latest(series["tc_import"]),
        "tcNorth": series_latest(series["tc_north"]),
        "tcSouth": series_latest(series["tc_south"]),
        "refinedOutput": series_latest(series["refined_output"]),
        "concentrateOutput": series_latest(series["concentrate_output"]),
        "galvanizedRate": series_latest(series["galvanized_rate"]),
        "zincOxideRate": series_latest(series["zinc_oxide_rate"]),
        "dieCastRate": series_latest(series["die_cast_rate"]),
        "shanghaiPremium": series_latest(series["shanghai_premium"]),
        "oreImport": series_latest(series["concentrate_import"]),
        "refinedNetImport": series_latest(refined_net_import),
        "monthSpread": series_latest(local["month_spread"]),
    }

    charts = {
        "shfePrice": seasonal_chart(series["price"]),
        "monthSpread": seasonal_chart(local["month_spread"]),
        "lmePrice": seasonal_chart(series["lme_price"]),
        "premium": continuous_chart(
            {
                "上海0#锌升贴水": series["shanghai_premium"],
                "广东升贴水（Excel）": local["premium_guangdong"],
                "天津升贴水（Excel）": local["premium_tianjin"],
            },
            limit=420,
        ),
        "tc": continuous_chart(
            {
                "北方50%Zn国产TC": series["tc_north"],
                "南方50%Zn国产TC": series["tc_south"],
                "进口50%Zn TC": series["tc_import"],
            },
            limit=420,
        ),
        "smeltingProfit": continuous_chart(
            {
                "国产矿冶炼利润（模型）": local["smelting_profit_domestic"],
                "进口矿冶炼利润（模型）": local["smelting_profit_import"],
            },
            limit=420,
        ),
        "refinedOutput": seasonal_chart(series["refined_output"]),
        "concentrateOutput": seasonal_chart(series["concentrate_output"]),
        "concentrateImport": seasonal_chart(series["concentrate_import"]),
        "refinedTrade": continuous_chart(
            {
                "进口": [[day, value / 10000] for day, value in series["refined_import"]],
                "出口": [[day, value / 10000] for day, value in series["refined_export"]],
            },
            limit=None,
        ),
        "refinedNetImport": seasonal_chart(refined_net_import),
        "shfeStock": seasonal_chart(series["shfe_stock"]),
        "lmeStock": seasonal_chart(series["lme_stock"]),
        "socialStock": seasonal_chart(series["social_stock"]),
        "concentratePortStock": seasonal_chart(series["concentrate_port_stock"]),
        "downstreamRates": continuous_chart(
            {
                "镀锌板卷开工率": series["galvanized_rate"],
                "氧化锌开工率": series["zinc_oxide_rate"],
                "锌合金开工率": series["die_cast_rate"],
            },
            limit=300,
        ),
        "galvanizedRateSeasonal": seasonal_chart(series["galvanized_rate"]),
        "zincOxideRateSeasonal": seasonal_chart(series["zinc_oxide_rate"]),
        "dieCastRateSeasonal": seasonal_chart(series["die_cast_rate"]),
        "galvanizedInventory": continuous_chart(
            {
                "镀锌板卷社会库存": series["galvanized_inventory"],
                "钢厂厂内库存": series["galvanized_mill_inventory"],
            },
            limit=300,
        ),
        "galvanizedInventorySeasonal": seasonal_chart(series["galvanized_inventory"]),
        "pmi": continuous_chart(
            {
                "中国制造业PMI新订单": local["pmi_china"],
                "美国ISM制造业PMI": local["pmi_us"],
            },
            limit=None,
        ),
    }

    stonex = local["stonex"]
    forecast_rows = stonex["forecast"]["rows"]
    metal_balance = metric_lookup(forecast_rows, "metal", "balance")
    concentrate_balance = metric_lookup(forecast_rows, "concentrate", "balance")
    refined_production = metric_lookup(forecast_rows, "refined", "production")
    consumption = metric_lookup(forecast_rows, "consumption")

    tc_value = series_latest(series["tc_import"])[1]
    profit_day, profit_value = series_latest(local["smelting_profit_domestic"])
    galvanized_value = series_latest(series["galvanized_rate"])[1]
    oxide_value = series_latest(series["zinc_oxide_rate"])[1]
    die_cast_value = series_latest(series["die_cast_rate"])[1]
    social_percentile = calendar_month_percentile(series["social_stock"])
    cycle_signals = [
        {
            "name": "矿端",
            "state": "极紧" if tc_value is not None and tc_value < 0 else "偏紧" if tc_value is not None and tc_value < 50 else "宽松",
            "detail": f"进口 TC {tc_value:,.1f} 美元/干吨" if tc_value is not None else "TC 暂无有效值",
            "tone": "down" if tc_value is not None and tc_value < 0 else "neutral",
        },
        {
            "name": "冶炼",
            "state": "亏损压产" if profit_value is not None and profit_value < 0 else "利润修复" if profit_value is not None else "待确认",
            "detail": f"国产矿模型利润 {profit_value:,.0f} 元/吨 · {profit_day}" if profit_value is not None else "利润模型暂无有效值",
            "tone": "down" if profit_value is not None and profit_value < 0 else "up" if profit_value is not None else "neutral",
        },
        {
            "name": "需求",
            "state": "镀锌偏强、其他分化" if galvanized_value is not None and galvanized_value >= 80 else "需求偏弱",
            "detail": f"镀锌 {galvanized_value or 0:.1f}% / 氧化锌 {oxide_value or 0:.1f}% / 合金 {die_cast_value or 0:.1f}%",
            "tone": "up" if galvanized_value is not None and galvanized_value >= 80 else "down",
        },
        {
            "name": "库存",
            "state": "季节性高位" if social_percentile is not None and social_percentile >= 75 else "季节性低位" if social_percentile is not None and social_percentile <= 25 else "季节性中位",
            "detail": f"国内现货库存处于历史同月 {social_percentile:.0f}% 分位" if social_percentile is not None else "历史同月分位待补",
            "tone": "down" if social_percentile is not None and social_percentile >= 75 else "up" if social_percentile is not None and social_percentile <= 25 else "neutral",
        },
    ]
    research_framework = [
        {"module": "定价", "metric": "沪锌 / LME 价格", "importance": "核心", "frequency": "日", "view": "季节性 + K线", "reason": "识别年度相对位置与趋势阶段"},
        {"module": "定价", "metric": "主连月差 / 三地升贴水", "importance": "核心", "frequency": "日", "view": "月差季节性；升贴水连续", "reason": "现货紧张最先反映在结构而非绝对价"},
        {"module": "矿端", "metric": "国内矿产量", "importance": "核心", "frequency": "月", "view": "季节性", "reason": "春节、检修和环保扰动强"},
        {"module": "矿端", "metric": "锌精矿进口", "importance": "核心", "frequency": "月", "view": "季节性", "reason": "国内原料缺口的主要补充项"},
        {"module": "矿端", "metric": "进口 / 国产 TC", "importance": "核心", "frequency": "周/日", "view": "连续趋势", "reason": "矿松紧与冶炼利润分配的核心价格"},
        {"module": "矿端", "metric": "港口库存", "importance": "高", "frequency": "周", "view": "季节性", "reason": "到港节奏具有明显日历效应"},
        {"module": "冶炼", "metric": "精炼锌产量", "importance": "核心", "frequency": "月", "view": "季节性", "reason": "检修、利润和原料共同决定供给"},
        {"module": "冶炼", "metric": "国产 / 进口矿冶炼利润", "importance": "核心", "frequency": "日", "view": "连续趋势", "reason": "判断减产兑现概率，不做季节性"},
        {"module": "冶炼", "metric": "精炼锌净进口", "importance": "高", "frequency": "月", "view": "季节性", "reason": "进口窗口与跨市场价差带来月度波动"},
        {"module": "需求", "metric": "镀锌 / 氧化锌 / 合金开工", "importance": "核心", "frequency": "周/月", "view": "分别做季节性", "reason": "混画会掩盖频率差异，春节效应显著"},
        {"module": "需求", "metric": "镀锌板卷库存", "importance": "高", "frequency": "周", "view": "季节性", "reason": "终端去库速度比单周开工更可靠"},
        {"module": "需求", "metric": "中国新订单 / 美国 ISM", "importance": "高", "frequency": "月", "view": "连续趋势 + 50线", "reason": "指标已季调，不再做季节性"},
        {"module": "库存", "metric": "SHFE / LME / 国内现货", "importance": "核心", "frequency": "日/周", "view": "季节性", "reason": "同月比较才能区分主动与被动累库"},
        {"module": "平衡", "metric": "国内矿端 / 锭端月度平衡", "importance": "核心", "frequency": "月", "view": "供需柱 + 平衡线", "reason": "残差与估算必须和实绩分层"},
        {"module": "全球", "metric": "StoneX 年度 / 季度平衡", "importance": "高", "frequency": "季/年", "view": "预测情景", "reason": "用于中期锚定，不与高频实绩混画"},
        {"module": "企业", "metric": "矿企 / 冶炼厂产量、指引、Capex", "importance": "高", "frequency": "季", "view": "同比 / 环比 / 指引差", "reason": "验证供给叙事是否在财报端兑现"},
    ]

    output = {
        "meta": {
            "title": "锌语新愿",
            "subtitle": "全球锌产业监测台",
            "builtAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "dataPriority": "直集 API > 最新 Excel 缓存值 > StoneX 2026年5月预测",
            "apiConnected": bool(key and any(api_series.values())),
            "warnings": warnings,
            "sourceFiles": [
                source_file_meta(INGOT_BOOK),
                source_file_meta(DEMAND_BOOK),
                source_file_meta(STONEX_BOOK),
                source_file_meta(COMPANY_BOOK),
            ],
        },
        "quote": quote,
        "latest": latest,
        "charts": charts,
        "kline": kline,
        "balances": {
            "mine": mine_balance,
            "refined": refined_balance,
            "methodology": [
                "矿端：国内锌精矿产量 + 进口矿实物量×45%，对比按精炼锌产量折算的矿耗。",
                "锭端：精炼锌产量 + 净进口构成供应；表观消费为扣除社会/保税库存变化后的残差。",
                "2026续接段优先使用直集，缺少锌合金净进口或保税库存时明确标注，不伪造补数。",
            ],
        },
        "stonex": stonex,
        "forecastSummary": {
            "years": stonex["forecast"]["years"],
            "metalBalance": metal_balance,
            "concentrateBalance": concentrate_balance,
            "refinedProduction": refined_production,
            "consumption": consumption,
        },
        "cycleSignals": cycle_signals,
        "researchFramework": research_framework,
        "companies": local["companies"],
        "projects": local["projects"],
        "sourceRegistry": api_meta,
        "calendar": [
            {"date": "2026-07-29", "company": "Glencore", "event": "H1产量报告 / 锌矿与冶炼锌"},
            {"date": "2026-07-29", "company": "Ivanhoe Mines", "event": "26Q2财报 / Kipushi"},
            {"date": "2026-08-05", "company": "Nexa", "event": "26Q2业绩 / 矿山与冶炼"},
            {"date": "2026-08-06", "company": "Korea Zinc", "event": "半年报（预计）"},
            {"date": "2026-08-07", "company": "Mitsui Mining", "event": "季度披露（预计）"},
            {"date": "2026-08-10", "company": "Peñoles", "event": "Torreon精炼锌（预计）"},
        ],
    }
    return output


def extract_reference_assets() -> None:
    if not TIN_REFERENCE.exists():
        raise FileNotFoundError(f"Tin reference is missing: {TIN_REFERENCE}")
    text = TIN_REFERENCE.read_text(encoding="utf-8")
    style_match = re.search(r"<style>(.*?)</style>", text, flags=re.S)
    if not style_match:
        raise RuntimeError("Unable to find the tin dashboard CSS")
    (ROOT / "styles.css").write_text(style_match.group(1).strip() + "\n", encoding="utf-8")

    scripts = re.findall(r"<script>(.*?)</script>", text, flags=re.S)
    chart_script = next(
        (
            script
            for script in scripts
            if "module.exports=e()" in script and "window.Chart=i" in script
        ),
        "",
    )
    if not chart_script:
        raise RuntimeError("Unable to find the embedded Chart.js distribution")
    vendor = ROOT / "vendor"
    vendor.mkdir(parents=True, exist_ok=True)
    (vendor / "chart.umd.min.js").write_text(chart_script.strip() + "\n", encoding="utf-8")


def write_data(data: dict[str, Any]) -> None:
    encoded = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    (ROOT / "data.js").write_text("window.ZINC_DATA=" + encoded + ";\n", encoding="utf-8")
    (ROOT / "data.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    quote_payload = {
        "updated_at": data["meta"]["builtAt"],
        "zn": data.get("quote") or {},
        "lme": {
            "last": data["latest"]["lme"][1],
            "date": data["latest"]["lme"][0],
            "source": "直集 API / 日度结算",
        },
    }
    (ROOT / "quotes.json").write_text(
        json.dumps(quote_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the Zinc Insight dashboard data.")
    parser.add_argument(
        "--skip-reference-assets",
        action="store_true",
        help="Do not re-extract CSS and Chart.js from the tin dashboard.",
    )
    args = parser.parse_args()
    if not args.skip_reference_assets:
        extract_reference_assets()
    data = build_data()
    write_data(data)
    password_hash = hashlib.sha256(b"yafco888").hexdigest()
    print(
        json.dumps(
            {
                "builtAt": data["meta"]["builtAt"],
                "apiConnected": data["meta"]["apiConnected"],
                "quote": data["quote"],
                "mineRows": len(data["balances"]["mine"]),
                "refinedRows": len(data["balances"]["refined"]),
                "mineCompanies": len(data["companies"]["mine"]["rows"]),
                "smelterCompanies": len(data["companies"]["smelter"]["rows"]),
                "passwordHash": password_hash,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
